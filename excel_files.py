# EXCEL FILES
# .XLSX
# openpyxl
import openpyxl

wb1 = openpyxl.load_workbook("october.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("november.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("december.xlsx", data_only=True)

# {"Pommes" : (760, 660, 900)}

def add_data_from_wb(wb, d):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        article_name = sheet.cell(row, 1).value
        if not article_name:
            break
        total_sales = sheet.cell(row, 4).value
        if d.get(article_name):
            d[article_name].append(total_sales)
        else:
            d[article_name] = [total_sales]

datas = {}
add_data_from_wb(wb1, datas)
add_data_from_wb(wb2, datas)
add_data_from_wb(wb3, datas)


print(datas)

wb_out = openpyxl.Workbook()
sheet = wb_out.active
sheet["A1"] = "ARTICLE"
sheet["B1"] = "OCTOBER"
sheet["C1"] = "NOVEMBER"
sheet["D1"] = "DECEMBER"

row = 2
for i in datas.items():
    # print(i)
    article_name = i[0]
    sales = i[1]
    sheet.cell(row, 1).value = article_name
    for j in range(0, len(sales)):
        sheet.cell(row, 2+j).value = sales[j]
    row += 1

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total sales in $")
chart = openpyxl.chart.BarChart3D()
chart.title = "Evolution of the price of apples in the last quarter"
chart.append(chart_serie)

sheet.add_chart(chart, "F2")

wb_out.save("total_sales_quarter.xlsx")



